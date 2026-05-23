# =============================================================================
# INSTALLATION:
#   pip install pandas numpy yfinance openpyxl tqdm requests requests_cache
# =============================================================================
#
# USAGE:
#   1. Create a CSV file named 'tickers.csv' with a column 'ticker'
#   2. Run: python major_cycle_analysis.py
#   3. Results saved to: output_results.xlsx and output_results.csv
#
# Pine Script: "Major Cycle: Pullback & Profit Analysis"
# Extended with: Fundamental metrics, sector, financial health score,
#                valuation zone (Buy/Hold), and overall buy rating (1-100).
# =============================================================================

import os
import time
import random
import logging
import warnings
import numpy as np
import pandas as pd
import yfinance as yf
import requests
import requests_cache
from tqdm import tqdm
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# =============================================================================
# CONFIGURATION
# =============================================================================

INPUT_CSV         = "tickers.csv"
OUTPUT_EXCEL      = "output_results.xlsx"
OUTPUT_CSV        = "output_results.csv"

LOOKBACK_PERIOD        = 252
PULLBACK_THRESHOLD     = -5.0
PROFIT_THRESHOLD       = 5.0
PIVOT_BARS             = 5

DOWNLOAD_PERIOD        = "max"   # Full history to match TradingView all-time accumulation
DOWNLOAD_INTERVAL      = "1d"

BATCH_SIZE             = 10
SLEEP_MIN              = 1.5
SLEEP_MAX              = 3.5
MAX_RETRIES            = 3
RETRY_BACKOFF_BASE     = 2

CACHE_EXPIRE_HOURS     = 6
LOG_FILE               = "major_cycle_analysis.log"

# =============================================================================
# SCORING WEIGHTS
# =============================================================================

# Financial Health Score sub-pillar weights (must sum to 100)
FH_WEIGHTS = {
    "profitability": 30,
    "balance_sheet": 25,
    "growth":        20,
    "cashflow":      15,
    "shareholder":   10,
}

# Overall Buy Rating pillar weights (must sum to 100)
RATING_WEIGHTS = {
    "financial_health": 40,
    "valuation_zone":   35,
    "momentum":         25,
}

# =============================================================================
# LOGGING
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =============================================================================
# REQUEST CACHING
# =============================================================================

requests_cache.install_cache(
    "yf_cache",
    expire_after=timedelta(hours=CACHE_EXPIRE_HOURS),
    allowable_methods=["GET"]
)

# =============================================================================
# HELPERS
# =============================================================================

def clamp_v(v, lo=0.0, hi=100.0):
    return max(lo, min(hi, v))


def safe(v):
    """Return None if NaN/None/inf, else round to 4dp float."""
    try:
        f = float(v)
        if np.isnan(f) or np.isinf(f):
            return None
        return round(f, 4)
    except Exception:
        return None


def pct(v):
    """Convert a ratio (0–1) to percentage safely."""
    r = safe(v)
    return round(r * 100, 4) if r is not None else None


# =============================================================================
# PINE SCRIPT REPLICATION
# =============================================================================

def ta_highest(series: pd.Series, length: int) -> pd.Series:
    """Replicates ta.highest(series, length)."""
    return series.rolling(window=length, min_periods=length).max()


def ta_lowest(series: pd.Series, length: int) -> pd.Series:
    """Replicates ta.lowest(series, length)."""
    return series.rolling(window=length, min_periods=length).min()


def ta_pivotlow(series: pd.Series, left_bars: int, right_bars: int) -> pd.Series:
    """
    Exact Pine Script ta.pivotlow replication.
    - Strict inequality both sides.
    - NaN on either side disqualifies the pivot.
    - Value placed at bar[i + right_bars] (confirmation bar).
    """
    arr = series.values
    n   = len(arr)
    out = np.full(n, np.nan)
    for i in range(left_bars, n - right_bars):
        val = arr[i]
        if np.isnan(val):
            continue
        if any(np.isnan(arr[i - j]) for j in range(1, left_bars + 1)):
            continue
        if any(np.isnan(arr[i + j]) for j in range(1, right_bars + 1)):
            continue
        if (all(arr[i - j] > val for j in range(1, left_bars + 1)) and
                all(arr[i + j] > val for j in range(1, right_bars + 1))):
            out[i + right_bars] = val
    return pd.Series(out, index=series.index)


def ta_pivothigh(series: pd.Series, left_bars: int, right_bars: int) -> pd.Series:
    """
    Exact Pine Script ta.pivothigh replication.
    - Strict inequality both sides.
    - Value placed at bar[i + right_bars] (confirmation bar).
    """
    arr = series.values
    n   = len(arr)
    out = np.full(n, np.nan)
    for i in range(left_bars, n - right_bars):
        val = arr[i]
        if np.isnan(val):
            continue
        if any(np.isnan(arr[i - j]) for j in range(1, left_bars + 1)):
            continue
        if any(np.isnan(arr[i + j]) for j in range(1, right_bars + 1)):
            continue
        if (all(arr[i - j] < val for j in range(1, left_bars + 1)) and
                all(arr[i + j] < val for j in range(1, right_bars + 1))):
            out[i + right_bars] = val
    return pd.Series(out, index=series.index)


# =============================================================================
# CYCLE INDICATOR  (Pine Script logic — exact translation)
# =============================================================================

def calculate_cycle_metrics(df: pd.DataFrame) -> dict:
    high  = df["High"]
    low   = df["Low"]
    close = df["Close"]

    # --- Drawdown (The Dip) ---
    rolling_ath   = ta_highest(high, LOOKBACK_PERIOD)
    drawdown_pct  = ((close - rolling_ath) / rolling_ath) * 100
    is_new_trough = ta_pivotlow(drawdown_pct, PIVOT_BARS, PIVOT_BARS)
    pullback_list = is_new_trough.dropna()
    pullback_list = pullback_list[pullback_list < PULLBACK_THRESHOLD].tolist()

    lower_bound      = min(pullback_list)             if pullback_list else np.nan
    typical_drawdown = float(np.mean(pullback_list))  if pullback_list else np.nan

    # --- Profit (The Rip) ---
    rolling_atl  = ta_lowest(low, LOOKBACK_PERIOD)
    profit_pct   = ((close - rolling_atl) / rolling_atl) * 100
    is_new_peak  = ta_pivothigh(profit_pct, PIVOT_BARS, PIVOT_BARS)
    profit_list  = is_new_peak.dropna()
    profit_list  = profit_list[profit_list > PROFIT_THRESHOLD].tolist()

    upper_bound    = max(profit_list)            if profit_list else np.nan
    typical_profit = float(np.mean(profit_list)) if profit_list else np.nan

    last_close    = float(close.iloc[-1])
    last_ath      = safe(rolling_ath.iloc[-1])
    last_atl      = safe(rolling_atl.iloc[-1])
    last_drawdown = safe(drawdown_pct.iloc[-1])
    last_profit   = safe(profit_pct.iloc[-1])
    last_date     = (df.index[-1].strftime("%Y-%m-%d")
                     if hasattr(df.index[-1], "strftime") else str(df.index[-1]))

    return {
        "current_close":          safe(last_close),
        "rolling_ath":            last_ath,
        "rolling_atl":            last_atl,
        "current_drawdown_pct":   last_drawdown,
        "current_profit_pct":     last_profit,
        "lower_bound":            safe(lower_bound),
        "upper_bound":            safe(upper_bound),
        "typical_drawdown":       safe(typical_drawdown),
        "typical_profit":         safe(typical_profit),
        "total_pullback_events":  len(pullback_list),
        "total_profit_events":    len(profit_list),
        "latest_bar_date":        last_date,
    }


# =============================================================================
# FUNDAMENTAL METRICS
# =============================================================================

def get_fundamentals(ticker_obj) -> dict:
    """Pull all fundamental metrics from yfinance .info."""
    try:
        info = ticker_obj.info
    except Exception:
        info = {}

    def g(key):
        return info.get(key, None)

    # Identity & sector
    sector   = g("sector")   or g("industryKey") or "N/A"
    industry = g("industry") or "N/A"
    name     = g("longName") or g("shortName")   or "N/A"
    country  = g("country")  or "N/A"
    exchange = g("exchange") or "N/A"

    # Market
    market_cap     = safe(g("marketCap"))
    enterprise_val = safe(g("enterpriseValue"))
    shares_out     = safe(g("sharesOutstanding"))
    float_shares   = safe(g("floatShares"))

    # Valuation
    trailing_pe    = safe(g("trailingPE"))
    forward_pe     = safe(g("forwardPE"))
    peg_ratio      = safe(g("pegRatio"))
    price_to_book  = safe(g("priceToBook"))
    price_to_sales = safe(g("priceToSalesTrailing12Months"))
    ev_to_ebitda   = safe(g("enterpriseToEbitda"))
    ev_to_revenue  = safe(g("enterpriseToRevenue"))

    # Profitability
    roe              = pct(g("returnOnEquity"))
    roa              = pct(g("returnOnAssets"))
    gross_margin     = pct(g("grossMargins"))
    operating_margin = pct(g("operatingMargins"))
    net_margin       = pct(g("profitMargins"))

    # Growth
    revenue_growth  = pct(g("revenueGrowth"))
    earnings_growth = pct(g("earningsGrowth"))
    eps_ttm         = safe(g("trailingEps"))
    eps_forward     = safe(g("forwardEps"))
    total_revenue   = safe(g("totalRevenue"))

    # Balance sheet
    total_debt    = safe(g("totalDebt"))
    total_cash    = safe(g("totalCash"))
    # yfinance returns D/E multiplied by 100 — normalise to ×1
    de_raw        = safe(g("debtToEquity"))
    debt_to_equity = round(de_raw / 100, 4) if de_raw is not None else None
    current_ratio  = safe(g("currentRatio"))
    quick_ratio    = safe(g("quickRatio"))

    # Cash flow
    free_cashflow  = safe(g("freeCashflow"))
    operating_cf   = safe(g("operatingCashflow"))
    ebitda         = safe(g("ebitda"))
    fcf_yield      = None
    fcf_margin     = None
    if free_cashflow and market_cap and market_cap > 0:
        fcf_yield  = round(free_cashflow / market_cap * 100, 4)
    if free_cashflow and total_revenue and total_revenue > 0:
        fcf_margin = round(free_cashflow / total_revenue * 100, 4)

    # Interest coverage (derived from financials — not directly in .info)
    interest_expense  = None
    interest_coverage = None
    try:
        fin = ticker_obj.financials
        if fin is not None and not fin.empty:
            ie_rows = [r for r in fin.index
                       if "interest" in r.lower() and "expense" in r.lower()]
            if ie_rows:
                ie_val = fin.loc[ie_rows[0]].iloc[0]
                interest_expense = safe(abs(ie_val))
                if ebitda and interest_expense and interest_expense > 0:
                    interest_coverage = round(ebitda / interest_expense, 4)
    except Exception:
        pass

    # Shareholder returns
    dividend_yield   = pct(g("dividendYield"))
    payout_ratio     = pct(g("payoutRatio"))
    short_pct_float  = pct(g("shortPercentOfFloat"))
    short_ratio      = safe(g("shortRatio"))

    # Share count change YoY (from balance sheet)
    shares_change_pct = None
    try:
        bs = ticker_obj.balance_sheet
        if bs is not None and not bs.empty and bs.shape[1] >= 2:
            so_rows = [r for r in bs.index if "share" in r.lower() and "issued" in r.lower()]
            if not so_rows:
                so_rows = [r for r in bs.index if "common stock" in r.lower()]
            if so_rows:
                s_now  = float(bs.loc[so_rows[0]].iloc[0])
                s_prev = float(bs.loc[so_rows[0]].iloc[1])
                if s_prev and s_prev != 0:
                    shares_change_pct = round((s_now - s_prev) / abs(s_prev) * 100, 4)
    except Exception:
        pass

    # Ownership
    insider_pct     = pct(g("heldPercentInsiders"))
    institution_pct = pct(g("heldPercentInstitutions"))

    # Analyst
    analyst_target  = safe(g("targetMeanPrice"))
    analyst_low     = safe(g("targetLowPrice"))
    analyst_high    = safe(g("targetHighPrice"))
    recommendation  = g("recommendationKey") or "N/A"
    num_analysts    = safe(g("numberOfAnalystOpinions"))

    # 52-week / technicals
    week52_high         = safe(g("fiftyTwoWeekHigh"))
    week52_low          = safe(g("fiftyTwoWeekLow"))
    week52_change       = pct(g("52WeekChange"))
    sp500_52wk_change   = pct(g("SandP52WeekChange"))
    rel_strength_vs_sp  = None
    if week52_change is not None and sp500_52wk_change is not None:
        rel_strength_vs_sp = round(week52_change - sp500_52wk_change, 4)
    beta = safe(g("beta"))

    return {
        # Identity
        "name":                       name,
        "sector":                     sector,
        "industry":                   industry,
        "country":                    country,
        "exchange":                   exchange,
        # Market
        "market_cap":                 market_cap,
        "enterprise_value":           enterprise_val,
        "shares_outstanding":         shares_out,
        "float_shares":               float_shares,
        # Valuation
        "trailing_pe":                trailing_pe,
        "forward_pe":                 forward_pe,
        "peg_ratio":                  peg_ratio,
        "price_to_book":              price_to_book,
        "price_to_sales":             price_to_sales,
        "ev_to_ebitda":               ev_to_ebitda,
        "ev_to_revenue":              ev_to_revenue,
        # Profitability
        "roe_pct":                    roe,
        "roa_pct":                    roa,
        "gross_margin_pct":           gross_margin,
        "operating_margin_pct":       operating_margin,
        "net_margin_pct":             net_margin,
        # Growth
        "revenue_growth_pct":         revenue_growth,
        "earnings_growth_pct":        earnings_growth,
        "eps_ttm":                    eps_ttm,
        "eps_forward":                eps_forward,
        "total_revenue":              total_revenue,
        # Balance Sheet
        "total_debt":                 total_debt,
        "total_cash":                 total_cash,
        "debt_to_equity":             debt_to_equity,
        "current_ratio":              current_ratio,
        "quick_ratio":                quick_ratio,
        "interest_coverage":          interest_coverage,
        # Cash Flow
        "free_cashflow":              free_cashflow,
        "operating_cashflow":         operating_cf,
        "fcf_yield_pct":              fcf_yield,
        "fcf_margin_pct":             fcf_margin,
        "ebitda":                     ebitda,
        # Shareholder Returns
        "dividend_yield_pct":         dividend_yield,
        "payout_ratio_pct":           payout_ratio,
        "shares_change_yoy_pct":      shares_change_pct,
        "short_pct_of_float":         short_pct_float,
        "short_ratio":                short_ratio,
        # Ownership
        "insider_ownership_pct":      insider_pct,
        "institution_ownership_pct":  institution_pct,
        # Analyst
        "analyst_target_price":       analyst_target,
        "analyst_low_price":          analyst_low,
        "analyst_high_price":         analyst_high,
        "analyst_recommendation":     recommendation,
        "num_analyst_opinions":       num_analysts,
        # Price / Technicals
        "week52_high":                week52_high,
        "week52_low":                 week52_low,
        "week52_change_pct":          week52_change,
        "sp500_52wk_change_pct":      sp500_52wk_change,
        "rel_strength_vs_sp500":      rel_strength_vs_sp,
        "beta":                       beta,
    }


# =============================================================================
# FINANCIAL HEALTH SCORE  (0–100, fundamentals only)
# =============================================================================

def score_financial_health(f: dict) -> tuple:
    """
    Five-pillar financial health score (0–100).

    Profitability (30%): ROE, gross/operating/net margins
    Balance Sheet (25%): D/E ratio, current ratio, interest coverage
    Growth        (20%): revenue growth, earnings growth
    Cash Flow     (15%): FCF yield, FCF margin
    Shareholder   (10%): payout sustainability, share dilution/buyback
    """
    scores = {}

    # ── Profitability ──────────────────────────────────────────────────────
    p = []
    if f["roe_pct"] is not None:
        r = f["roe_pct"]
        p.append(clamp_v(100 if r >= 20 else 80 if r >= 15 else 60 if r >= 10
                         else 40 if r >= 5 else 20 if r >= 0 else 0))
    if f["gross_margin_pct"] is not None:
        gm = f["gross_margin_pct"]
        p.append(clamp_v(100 if gm >= 50 else 85 if gm >= 40 else 65 if gm >= 30
                          else 45 if gm >= 20 else 20))
    if f["operating_margin_pct"] is not None:
        om = f["operating_margin_pct"]
        p.append(clamp_v(100 if om >= 20 else 80 if om >= 15 else 60 if om >= 10
                          else 40 if om >= 5 else 20 if om >= 0 else 0))
    if f["net_margin_pct"] is not None:
        nm = f["net_margin_pct"]
        p.append(clamp_v(100 if nm >= 15 else 80 if nm >= 10 else 60 if nm >= 5
                          else 20 if nm >= 0 else 0))
    scores["profitability"] = float(np.mean(p)) if p else 50.0

    # ── Balance Sheet ──────────────────────────────────────────────────────
    bs = []
    if f["debt_to_equity"] is not None:
        de = f["debt_to_equity"]
        bs.append(clamp_v(100 if de < 0.3 else 85 if de < 0.5 else 65 if de < 1.0
                           else 40 if de < 2.0 else 15))
    if f["current_ratio"] is not None:
        cr = f["current_ratio"]
        bs.append(clamp_v(100 if cr >= 2.0 else 80 if cr >= 1.5 else 60 if cr >= 1.2
                           else 40 if cr >= 1.0 else 10))
    if f["interest_coverage"] is not None:
        ic = f["interest_coverage"]
        bs.append(clamp_v(100 if ic >= 10 else 80 if ic >= 5 else 60 if ic >= 3
                           else 35 if ic >= 1.5 else 5))
    scores["balance_sheet"] = float(np.mean(bs)) if bs else 50.0

    # ── Growth ────────────────────────────────────────────────────────────
    gr = []
    if f["revenue_growth_pct"] is not None:
        rg = f["revenue_growth_pct"]
        gr.append(clamp_v(100 if rg >= 20 else 85 if rg >= 15 else 70 if rg >= 10
                           else 50 if rg >= 5 else 30 if rg >= 0 else 10))
    if f["earnings_growth_pct"] is not None:
        eg = f["earnings_growth_pct"]
        gr.append(clamp_v(100 if eg >= 25 else 85 if eg >= 15 else 65 if eg >= 5
                           else 40 if eg >= 0 else 10))
    scores["growth"] = float(np.mean(gr)) if gr else 50.0

    # ── Cash Flow ─────────────────────────────────────────────────────────
    cf = []
    if f["fcf_yield_pct"] is not None:
        fy = f["fcf_yield_pct"]
        cf.append(clamp_v(100 if fy >= 6 else 80 if fy >= 4 else 60 if fy >= 2
                           else 35 if fy >= 0 else 5))
    if f["fcf_margin_pct"] is not None:
        fm = f["fcf_margin_pct"]
        cf.append(clamp_v(100 if fm >= 20 else 85 if fm >= 15 else 70 if fm >= 10
                           else 50 if fm >= 5 else 25 if fm >= 0 else 0))
    scores["cashflow"] = float(np.mean(cf)) if cf else 50.0

    # ── Shareholder ───────────────────────────────────────────────────────
    sh = []
    # Payout ratio: very high = risky; no dividend = neutral (company reinvesting)
    if f["payout_ratio_pct"] is not None:
        pr = f["payout_ratio_pct"]
        sh.append(clamp_v(100 if pr < 40 else 75 if pr < 60 else 45 if pr < 80 else 15))
    else:
        sh.append(60.0)   # no dividend — reinvesting, neutral positive signal
    # Share dilution: buybacks = 100, flat = 70, mild dilution <3% = 50, heavy = 0
    if f["shares_change_yoy_pct"] is not None:
        sc = f["shares_change_yoy_pct"]
        sh.append(clamp_v(100 if sc < -2 else 70 if sc < 0 else 50 if sc < 3
                           else 25 if sc < 10 else 0))
    scores["shareholder"] = float(np.mean(sh)) if sh else 50.0

    # ── Weighted total ────────────────────────────────────────────────────
    total = round(clamp_v(
        sum(scores[k] * FH_WEIGHTS[k] / 100 for k in FH_WEIGHTS)
    ), 1)

    if total >= 80:   label = "Excellent"
    elif total >= 65: label = "Good"
    elif total >= 50: label = "Fair"
    elif total >= 35: label = "Weak"
    else:             label = "Poor"

    return total, label


# =============================================================================
# VALUATION ZONE  (cycle drawdown metrics)
# =============================================================================

def calculate_valuation_zone(cycle: dict) -> tuple:
    """
    Determines valuation zone and score (0–100) from cycle metrics.

    STRONG BUY : current_drawdown <= typical_drawdown (historically deep dip)
    BUY        : typical_drawdown < drawdown <= 0.5 * typical_drawdown
    WATCH      : 0.5 * typical_drawdown < drawdown <= -5%
    HOLD       : drawdown > -5% (near all-time highs, limited safety margin)

    Valuation score:
      100  = at or below lower_bound (maximum historical dip)
      70–100 = between typical_drawdown and lower_bound
      40–70  = between 0.5x typical and typical_drawdown
      10–40  = mild pullback (-5% to 0.5x typical)
      0–10   = near all-time high
    """
    dd = cycle.get("current_drawdown_pct")
    td = cycle.get("typical_drawdown")
    lb = cycle.get("lower_bound")

    if dd is None:
        return "Insufficient Data", 0.0

    # Valuation score calculation
    if td is not None and lb is not None and td < 0 and lb < 0:
        if dd <= lb:
            val_score = 100.0
        elif dd <= td:
            # Linear 70→100 from td to lb
            span = lb - td
            val_score = 70.0 + 30.0 * (dd - td) / (span if span != 0 else -1e-9)
        elif dd <= td * 0.5:
            # Linear 40→70 from 0.5*td to td
            half_td = td * 0.5
            span    = td - half_td
            val_score = 40.0 + 30.0 * (dd - half_td) / (span if span != 0 else -1e-9)
        elif dd <= -5.0:
            # Linear 10→40 from -5% to 0.5*td
            half_td = td * 0.5
            span    = half_td - (-5.0)
            val_score = 10.0 + 30.0 * (dd - (-5.0)) / (span if span != 0 else -1e-9)
        else:
            # Near highs: 0–10
            val_score = max(0.0, 10.0 + dd * 2.0)
    else:
        # No pivot history — simple heuristic on drawdown alone (max 60)
        val_score = clamp_v(-dd * 2.0, 0.0, 60.0)

    val_score = round(clamp_v(val_score, 0.0, 100.0), 1)

    # Zone label
    if td is not None and td < 0:
        if dd <= td:
            zone = "STRONG BUY"
        elif dd <= td * 0.5:
            zone = "BUY"
        elif dd <= -5.0:
            zone = "WATCH"
        else:
            zone = "HOLD"
    else:
        # Fallback when no pivot history available
        zone = ("STRONG BUY" if dd <= -20 else "BUY" if dd <= -10
                else "WATCH" if dd <= -5 else "HOLD")

    return zone, val_score


# =============================================================================
# OVERALL BUY RATING  (0–100)
# =============================================================================

def calculate_overall_rating(fh_score: float, val_score: float, cycle: dict) -> tuple:
    """
    Three-pillar weighted rating:
      Financial Health (40%) : business quality & durability
      Valuation Zone   (35%) : price vs historical dip opportunity
      Momentum         (25%) : signal reliability & reward/risk of cycle

    Momentum sub-score:
      - Events score  (50%): number of historical pullback + profit events
        (calibration quality — more events = more reliable signal)
      - R/R score     (50%): typical_profit / |typical_drawdown|
        (reward/risk — want ratio > 1.5; 3x = perfect)
    """
    pull_events = cycle.get("total_pullback_events") or 0
    prof_events = cycle.get("total_profit_events")   or 0
    typ_dd      = cycle.get("typical_drawdown")
    typ_pr      = cycle.get("typical_profit")

    # Events calibration score (20 combined events = full score)
    events_score = clamp_v((pull_events + prof_events) / 20.0 * 100.0, 0.0, 100.0)

    # Reward/risk score
    rr_score = 50.0
    if typ_dd and typ_pr and typ_dd < 0 and typ_pr > 0:
        rr = typ_pr / abs(typ_dd)
        rr_score = clamp_v(rr / 3.0 * 100.0, 0.0, 100.0)

    momentum_score = round(events_score * 0.5 + rr_score * 0.5, 1)

    raw = (fh_score       * RATING_WEIGHTS["financial_health"] / 100.0 +
           val_score      * RATING_WEIGHTS["valuation_zone"]   / 100.0 +
           momentum_score * RATING_WEIGHTS["momentum"]         / 100.0)

    rating = int(round(clamp_v(raw, 0.0, 100.0)))

    if rating >= 80:   label = "STRONG BUY"
    elif rating >= 65: label = "BUY"
    elif rating >= 50: label = "WATCH"
    elif rating >= 35: label = "HOLD"
    else:              label = "AVOID"

    return rating, label


# =============================================================================
# DATA DOWNLOAD
# =============================================================================

def download_yfinance(ticker_str: str):
    """Returns (df_ohlcv, ticker_obj) or (None, None)."""
    try:
        t  = yf.Ticker(ticker_str)
        df = t.history(period=DOWNLOAD_PERIOD, interval=DOWNLOAD_INTERVAL, auto_adjust=True)
        if df is None or df.empty:
            return None, None
        df.index = pd.to_datetime(df.index)
        if df.index.tz is not None:
            df.index = df.index.tz_convert(None)
        df = df[["Open", "High", "Low", "Close", "Volume"]]
        # FIX #8: only drop rows where Close is missing; ffill others
        df = df[df["Close"].notna()]
        df[["Open", "High", "Low", "Volume"]] = (
            df[["Open", "High", "Low", "Volume"]].ffill()
        )
        if len(df) < LOOKBACK_PERIOD + PIVOT_BARS * 2 + 10:
            return None, None
        return df, t
    except Exception as e:
        logger.debug(f"yfinance error for {ticker_str}: {e}")
        return None, None


def download_stooq(ticker_str: str):
    """Stooq fallback — returns (df_ohlcv, None). Full history, no cutoff."""
    try:
        url  = f"https://stooq.com/q/d/l/?s={ticker_str.lower()}.us&i=d"
        resp = requests.get(url, timeout=15)
        if resp.status_code != 200 or len(resp.text) < 100:
            return None, None
        from io import StringIO
        df = pd.read_csv(StringIO(resp.text))
        df.columns = [c.strip().capitalize() for c in df.columns]
        if "Date" not in df.columns:
            return None, None
        df["Date"] = pd.to_datetime(df["Date"])
        df.set_index("Date", inplace=True)
        df.sort_index(inplace=True)
        for col in ["Open", "High", "Low", "Close", "Volume"]:
            if col not in df.columns:
                return None, None
        df = df[["Open", "High", "Low", "Close", "Volume"]]
        # FIX #7 + #8: no hardcoded cutoff; preserve bar index
        df = df[df["Close"].notna()]
        df[["Open", "High", "Low", "Volume"]] = (
            df[["Open", "High", "Low", "Volume"]].ffill()
        )
        if len(df) < LOOKBACK_PERIOD + PIVOT_BARS * 2 + 10:
            return None, None
        return df, None
    except Exception as e:
        logger.debug(f"stooq error for {ticker_str}: {e}")
        return None, None


def download_with_retry(ticker_str: str):
    """Try yfinance with retries, then stooq fallback."""
    for attempt in range(1, MAX_RETRIES + 1):
        df, t = download_yfinance(ticker_str)
        if df is not None:
            return df, t
        sleep_time = RETRY_BACKOFF_BASE ** attempt + random.uniform(0, 1)
        logger.debug(f"{ticker_str}: attempt {attempt} failed — retry in {sleep_time:.1f}s")
        time.sleep(sleep_time)

    logger.debug(f"{ticker_str}: trying stooq fallback")
    df, t = download_stooq(ticker_str)
    if df is not None:
        logger.info(f"{ticker_str}: using stooq data (Yahoo failed)")
        return df, t

    logger.warning(f"{ticker_str}: all data sources failed")
    return None, None


# =============================================================================
# COLUMN DEFINITIONS  (order for output)
# =============================================================================

ID_COLS      = ["ticker", "name", "sector", "industry", "country", "exchange"]
SCORE_COLS   = ["overall_rating", "overall_rating_label",
                "financial_health_score", "financial_health_label",
                "valuation_zone", "valuation_score"]
CYCLE_COLS   = ["current_close", "rolling_ath", "rolling_atl",
                "current_drawdown_pct", "current_profit_pct",
                "lower_bound", "upper_bound", "typical_drawdown", "typical_profit",
                "total_pullback_events", "total_profit_events", "latest_bar_date"]
MKT_COLS     = ["market_cap", "enterprise_value", "shares_outstanding",
                "float_shares", "beta"]
VAL_COLS     = ["trailing_pe", "forward_pe", "peg_ratio", "price_to_book",
                "price_to_sales", "ev_to_ebitda", "ev_to_revenue"]
PROF_COLS    = ["roe_pct", "roa_pct", "gross_margin_pct",
                "operating_margin_pct", "net_margin_pct"]
GROWTH_COLS  = ["revenue_growth_pct", "earnings_growth_pct",
                "eps_ttm", "eps_forward", "total_revenue"]
BS_COLS      = ["total_debt", "total_cash", "debt_to_equity",
                "current_ratio", "quick_ratio", "interest_coverage"]
CF_COLS      = ["free_cashflow", "operating_cashflow",
                "fcf_yield_pct", "fcf_margin_pct", "ebitda"]
SH_COLS      = ["dividend_yield_pct", "payout_ratio_pct", "shares_change_yoy_pct",
                "short_pct_of_float", "short_ratio"]
OWN_COLS     = ["insider_ownership_pct", "institution_ownership_pct"]
ANALYST_COLS = ["analyst_target_price", "analyst_low_price", "analyst_high_price",
                "analyst_recommendation", "num_analyst_opinions"]
TECH_COLS    = ["week52_high", "week52_low", "week52_change_pct",
                "sp500_52wk_change_pct", "rel_strength_vs_sp500"]

ALL_COLS = (ID_COLS + SCORE_COLS + CYCLE_COLS + MKT_COLS + VAL_COLS +
            PROF_COLS + GROWTH_COLS + BS_COLS + CF_COLS + SH_COLS +
            OWN_COLS + ANALYST_COLS + TECH_COLS)

FUND_KEYS = [c for c in ALL_COLS if c not in (ID_COLS[:1] + SCORE_COLS + CYCLE_COLS)]


def empty_row(ticker_str: str) -> dict:
    row = {k: None for k in ALL_COLS}
    row["ticker"] = ticker_str
    return row


# =============================================================================
# BATCH PROCESSING
# =============================================================================

def process_tickers(tickers: list) -> pd.DataFrame:
    results = []
    failed  = []
    batches = [tickers[i:i + BATCH_SIZE] for i in range(0, len(tickers), BATCH_SIZE)]
    logger.info(f"Processing {len(tickers)} tickers in {len(batches)} batches of {BATCH_SIZE}")

    with tqdm(total=len(tickers), desc="Analysing tickers", unit="ticker") as pbar:
        for batch_idx, batch in enumerate(batches):
            for ticker_str in batch:
                try:
                    df, ticker_obj = download_with_retry(ticker_str)
                    if df is None:
                        failed.append(ticker_str)
                        results.append(empty_row(ticker_str))
                        pbar.update(1)
                        continue

                    cycle = calculate_cycle_metrics(df)

                    if ticker_obj is not None:
                        fund = get_fundamentals(ticker_obj)
                    else:
                        fund = {k: None for k in FUND_KEYS}
                        fund.update({"name": "N/A", "sector": "N/A", "industry": "N/A",
                                     "country": "N/A", "exchange": "N/A"})

                    fh_score,   fh_label     = score_financial_health(fund)
                    val_zone,   val_score    = calculate_valuation_zone(cycle)
                    rating,     rating_label = calculate_overall_rating(
                                                    fh_score, val_score, cycle)

                    row = {"ticker": ticker_str}
                    row.update(cycle)
                    row.update(fund)
                    row["financial_health_score"] = fh_score
                    row["financial_health_label"] = fh_label
                    row["valuation_zone"]         = val_zone
                    row["valuation_score"]        = val_score
                    row["overall_rating"]         = rating
                    row["overall_rating_label"]   = rating_label

                    results.append(row)
                    logger.info(
                        f"{ticker_str:6s} | {str(fund.get('sector','?')):22s} | "
                        f"FH={fh_score:5.1f} ({fh_label:9s}) | "
                        f"Zone={val_zone:11s} | Rating={rating:3d}/100 [{rating_label}]"
                    )

                except Exception as e:
                    logger.error(f"{ticker_str}: unexpected error: {e}", exc_info=True)
                    failed.append(ticker_str)
                    results.append(empty_row(ticker_str))

                pbar.update(1)

            if batch_idx < len(batches) - 1:
                time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

    if failed:
        logger.warning(f"Failed tickers ({len(failed)}): {', '.join(failed)}")

    return pd.DataFrame(results)


# =============================================================================
# EXCEL FORMATTING
# =============================================================================

# =============================================================================
# COLOUR PALETTE
# =============================================================================
C_STRONG_BUY  = "006400"   # Strong Buy / Excellent — dark green
C_BUY         = "228B22"   # Buy / Good             — forest green
C_WATCH       = "FFCC00"   # Watch / Fair           — gold
C_HOLD        = "FF4500"   # Hold / Weak            — orange-red
C_AVOID       = "B22222"   # Avoid / Poor           — firebrick red
C_HEADER      = "1A1A1B"   # Header + section dividers — near black
C_GREY_ALT    = "F5F7FA"   # Alternating row tint
C_WHITE       = "FFFFFF"
C_TEXT_WHITE  = "FFFFFF"   # All text on coloured backgrounds
C_TEXT_BLACK  = "000000"   # Text on light/white backgrounds

# Aliases so existing colour_numeric / colour_label calls stay readable
C_DARK_GREEN  = C_STRONG_BUY
C_MID_GREEN   = C_BUY
C_AMBER       = C_WATCH
C_ORANGE      = C_HOLD
C_RED_LIGHT   = C_AVOID
C_BLUE_HDR    = C_HEADER
# C_LIGHT_GREEN is no longer used — Watch/Fair now uses C_WATCH (gold)


def mk_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def mk_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)


def mk_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_sheet_formatting(ws, df_cols: list):
    """
    Three-pass formatting — order matters:
      Pass 1: Header row  (no wrap_text)
      Pass 2: Alternating row fills — skips colour-coded columns entirely
              so the colour-coding pass can never be overwritten
      Pass 3: Colour-coded columns applied cleanly on top
    """

    # ── Pass 1: Header ────────────────────────────────────────────────────
    hdr_fill  = mk_fill(C_BLUE_HDR)
    hdr_font  = mk_font(bold=True, color=C_TEXT_WHITE, size=10)
    # wrap_text=False — user explicitly does not want header text wrapping
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = mk_border()
    ws.row_dimensions[1].height = 20   # single-line height since no wrapping

    # Build column-name → column-letter map from header row
    col_map = {cell.value: get_column_letter(cell.column) for cell in ws[1]}

    # Define which columns will receive colour-coding in Pass 3.
    # These columns are deliberately SKIPPED in Pass 2 (alternating rows)
    # so alternating fills never paint over the colours we apply in Pass 3.
    COLOUR_CODED_COLS = {
        "overall_rating", "overall_rating_label",
        "financial_health_score", "financial_health_label",
        "valuation_zone",
        "current_drawdown_pct",
        "debt_to_equity",
        "peg_ratio",
        "roe_pct",
        "fcf_yield_pct",
    }
    # Build the set of column letters that are colour-coded
    colour_coded_letters = {
        col_map[name] for name in COLOUR_CODED_COLS if name in col_map
    }

    # ── Pass 2: Alternating row fills (skipping colour-coded columns) ─────
    for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        alt_fill = mk_fill(C_GREY_ALT) if ri % 2 == 0 else mk_fill(C_WHITE)
        for cell in row:
            cell.border    = mk_border()
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=False)
            # Only apply alternating fill to non-colour-coded columns
            if get_column_letter(cell.column) not in colour_coded_letters:
                cell.fill = alt_fill

    # ── Pass 3: Colour-coding ─────────────────────────────────────────────
    # Helper: apply threshold rules to a numeric column
    # rules = list of (condition_fn, bg_hex, fg_hex, bold) — first match wins
    def colour_numeric(col_name, rules):
        if col_name not in col_map:
            return
        cl = col_map[col_name]
        for ri in range(2, ws.max_row + 1):
            cell = ws[f"{cl}{ri}"]
            v    = cell.value
            if v is None:
                continue
            try:
                v = float(v)
            except (TypeError, ValueError):
                continue
            for cond, bg, fg, bold in rules:
                if cond(v):
                    cell.fill = mk_fill(bg)
                    cell.font = mk_font(bold=bold, color=fg)
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center", wrap_text=False)
                    break

    # Helper: apply lookup map to a text/label column
    def colour_label(col_name, label_map):
        if col_name not in col_map:
            return
        cl = col_map[col_name]
        for ri in range(2, ws.max_row + 1):
            cell = ws[f"{cl}{ri}"]
            v    = str(cell.value).strip() if cell.value is not None else ""
            if v in label_map:
                bg, fg, bold = label_map[v]
                cell.fill = mk_fill(bg)
                cell.font = mk_font(bold=bold, color=fg)
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center", wrap_text=False)

    # ── overall_rating  (numeric 0–100) ───────────────────────────────────
    colour_numeric("overall_rating", [
        (lambda v: v >= 80, C_STRONG_BUY, C_TEXT_WHITE, True),
        (lambda v: v >= 65, C_BUY,        C_TEXT_WHITE, True),
        (lambda v: v >= 50, C_WATCH,      C_TEXT_WHITE, True),
        (lambda v: v >= 35, C_HOLD,       C_TEXT_WHITE, True),
        (lambda v: True,    C_AVOID,      C_TEXT_WHITE, True),
    ])

    # ── overall_rating_label  (text) ──────────────────────────────────────
    colour_label("overall_rating_label", {
        "STRONG BUY": (C_STRONG_BUY, C_TEXT_WHITE, True),
        "BUY":        (C_BUY,        C_TEXT_WHITE, True),
        "WATCH":      (C_WATCH,      C_TEXT_WHITE, True),
        "HOLD":       (C_HOLD,       C_TEXT_WHITE, True),
        "AVOID":      (C_AVOID,      C_TEXT_WHITE, True),
    })

    # ── financial_health_score  (numeric 0–100) ───────────────────────────
    colour_numeric("financial_health_score", [
        (lambda v: v >= 80, C_STRONG_BUY, C_TEXT_WHITE, True),
        (lambda v: v >= 65, C_BUY,        C_TEXT_WHITE, True),
        (lambda v: v >= 50, C_WATCH,      C_TEXT_WHITE, True),
        (lambda v: v >= 35, C_HOLD,       C_TEXT_WHITE, True),
        (lambda v: True,    C_AVOID,      C_TEXT_WHITE, True),
    ])

    # ── financial_health_label  (text) ────────────────────────────────────
    colour_label("financial_health_label", {
        "Excellent": (C_STRONG_BUY, C_TEXT_WHITE, True),
        "Good":      (C_BUY,        C_TEXT_WHITE, True),
        "Fair":      (C_WATCH,      C_TEXT_WHITE, True),
        "Weak":      (C_HOLD,       C_TEXT_WHITE, True),
        "Poor":      (C_AVOID,      C_TEXT_WHITE, True),
    })

    # ── valuation_zone  (text) ────────────────────────────────────────────
    colour_label("valuation_zone", {
        "STRONG BUY":        (C_STRONG_BUY, C_TEXT_WHITE, True),
        "BUY":               (C_BUY,        C_TEXT_WHITE, True),
        "WATCH":             (C_WATCH,      C_TEXT_WHITE, True),
        "HOLD":              (C_HOLD,       C_TEXT_WHITE, True),
        "Insufficient Data": (C_GREY_ALT,   C_TEXT_BLACK, False),
    })

    # ── current_drawdown_pct  (deeper dip = greener) ──────────────────────
    colour_numeric("current_drawdown_pct", [
        (lambda v: v <= -20, C_STRONG_BUY, C_TEXT_WHITE, True),
        (lambda v: v <= -10, C_BUY,        C_TEXT_WHITE, True),
        (lambda v: v <= -5,  C_WATCH,      C_TEXT_WHITE, True),
        (lambda v: v <= -2,  C_HOLD,       C_TEXT_WHITE, True),
        (lambda v: True,     C_AVOID,      C_TEXT_WHITE, True),
    ])

    # ── debt_to_equity  (lower = greener) ─────────────────────────────────
    colour_numeric("debt_to_equity", [
        (lambda v: v < 0.3, C_STRONG_BUY, C_TEXT_WHITE, True),
        (lambda v: v < 0.5, C_BUY,        C_TEXT_WHITE, True),
        (lambda v: v < 1.0, C_WATCH,      C_TEXT_WHITE, True),
        (lambda v: v < 2.0, C_HOLD,       C_TEXT_WHITE, True),
        (lambda v: True,    C_AVOID,      C_TEXT_WHITE, True),
    ])

    # ── peg_ratio  (lower = greener; ignore negatives — no earnings) ──────
    colour_numeric("peg_ratio", [
        (lambda v: v <= 0,  C_GREY_ALT,   C_TEXT_BLACK, False),
        (lambda v: v < 1.0, C_STRONG_BUY, C_TEXT_WHITE, True),
        (lambda v: v < 1.5, C_BUY,        C_TEXT_WHITE, True),
        (lambda v: v < 2.5, C_WATCH,      C_TEXT_WHITE, True),
        (lambda v: v < 4.0, C_HOLD,       C_TEXT_WHITE, True),
        (lambda v: True,    C_AVOID,      C_TEXT_WHITE, True),
    ])

    # ── roe_pct  (higher = greener) ───────────────────────────────────────
    colour_numeric("roe_pct", [
        (lambda v: v >= 20, C_STRONG_BUY, C_TEXT_WHITE, True),
        (lambda v: v >= 15, C_BUY,        C_TEXT_WHITE, True),
        (lambda v: v >= 10, C_WATCH,      C_TEXT_WHITE, True),
        (lambda v: v >= 5,  C_HOLD,       C_TEXT_WHITE, True),
        (lambda v: True,    C_AVOID,      C_TEXT_WHITE, True),
    ])

    # ── fcf_yield_pct  (higher = greener) ────────────────────────────────
    colour_numeric("fcf_yield_pct", [
        (lambda v: v >= 6,  C_STRONG_BUY, C_TEXT_WHITE, True),
        (lambda v: v >= 4,  C_BUY,        C_TEXT_WHITE, True),
        (lambda v: v >= 2,  C_WATCH,      C_TEXT_WHITE, True),
        (lambda v: v >= 0,  C_HOLD,       C_TEXT_WHITE, True),
        (lambda v: True,    C_AVOID,      C_TEXT_WHITE, True),
    ])

    # ── Autofit columns (content-driven, min 10, no hard cap) ────────────
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = max(max_len + 3, 10)

    # ── Autofit rows (content-driven, min 15) ────────────────────────────
    for ri in range(2, ws.max_row + 1):
        row_cells  = list(ws.iter_rows(min_row=ri, max_row=ri))[0]
        max_chars  = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in row_cells
        )
        # Standard single-line height; only expand if content is genuinely long
        ws.row_dimensions[ri].height = 15 if max_chars < 60 else 28

    ws.freeze_panes = "C2"


def add_summary_sheet(wb, df: pd.DataFrame):
    """
    Sorted top-picks sheet with key metrics only.
    Colour-coding applied via apply_sheet_formatting after all data is written,
    so there is no overwrite conflict.
    """
    ws = wb.create_sheet("Summary — Top Picks")

    summary_cols = [
        "ticker", "name", "sector",
        "overall_rating", "overall_rating_label",
        "financial_health_score", "financial_health_label",
        "valuation_zone", "valuation_score",
        "current_drawdown_pct", "typical_drawdown", "lower_bound",
        "current_profit_pct", "typical_profit",
        "trailing_pe", "peg_ratio",
        "roe_pct", "fcf_yield_pct", "debt_to_equity",
        "revenue_growth_pct", "earnings_growth_pct",
        "net_margin_pct", "gross_margin_pct",
        "analyst_recommendation", "analyst_target_price",
        "latest_bar_date",
    ]
    available = [c for c in summary_cols if c in df.columns]
    sub = df[available].copy().sort_values(
        "overall_rating", ascending=False, na_position="last"
    )

    # ── Write header (no wrap_text) ───────────────────────────────────────
    hdr_fill  = mk_fill(C_BLUE_HDR)
    hdr_font  = mk_font(bold=True, color=C_TEXT_WHITE, size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for ci, col in enumerate(available, start=1):
        cell            = ws.cell(row=1, column=ci, value=col)
        cell.fill       = hdr_fill
        cell.font       = hdr_font
        cell.alignment  = hdr_align
        cell.border     = mk_border()
    ws.row_dimensions[1].height = 20

    # ── Write data rows with plain alternating fills ──────────────────────
    # Colour-coded columns are written with a neutral fill here;
    # apply_sheet_formatting (Pass 3) will paint them correctly afterwards.
    for ri, (_, row_data) in enumerate(sub.iterrows(), start=2):
        alt_fill = mk_fill(C_GREY_ALT) if ri % 2 == 0 else mk_fill(C_WHITE)
        for ci, col in enumerate(available, start=1):
            cell           = ws.cell(row=ri, column=ci, value=row_data[col])
            cell.fill      = alt_fill
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=False)
            cell.border    = mk_border()

    # ── Apply full formatting (Pass 2 will skip colour-coded cols,
    #    Pass 3 will paint them — no overwrite conflict) ───────────────────
    apply_sheet_formatting(ws, available)

    ws.freeze_panes = "C2"


def add_legend_sheet(wb):
    """
    Scoring legend sheet.
    - Header row          : C_HEADER (1A1A1B), White Bold, no wrap text
    - Section dividers    : C_HEADER (1A1A1B), White Bold, no wrap text
                            (matches header style exactly as requested)
    - Data rows           : white background, normal text, no wrap text
    - Autofit columns and rows based on content
    """
    ws = wb.create_sheet("Legend & Scoring Guide")

    std_align  = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ctr_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    hdr_fill   = mk_fill(C_HEADER)
    hdr_font   = mk_font(bold=True, color=C_TEXT_WHITE, size=10)
    data_fill  = mk_fill(C_WHITE)

    rows = [
        # (col_A,                col_B,                     col_C)
        ("SECTION",              "METRIC / LABEL",           "WHAT IT MEANS & THRESHOLDS"),
        ("── OVERALL RATING ──", "80–100  STRONG BUY",      "Excellent fundamentals + deep dip. High-conviction buy."),
        ("",                     "65–79   BUY",              "Good quality business entering buy zone."),
        ("",                     "50–64   WATCH",            "Decent business but price not yet attractive."),
        ("",                     "35–49   HOLD",             "Weak fundamentals or near all-time high."),
        ("",                     "0–34    AVOID",            "Poor fundamentals and/or overvalued."),
        ("── FINANCIAL HEALTH ──","Excellent (80–100)",      "Strong across all 5 pillars."),
        ("",                     "Good (65–79)",             "Solid business, minor weaknesses."),
        ("",                     "Fair (50–64)",             "Mixed — some areas of concern."),
        ("",                     "Weak (35–49)",             "Multiple red flags."),
        ("",                     "Poor (0–34)",              "Serious fundamental problems."),
        ("── VALUATION ZONE ──", "STRONG BUY",               "Current drawdown <= typical historical dip. Historically cheap."),
        ("",                     "BUY",                      "Drawdown between 50-100% of typical dip. Approaching value."),
        ("",                     "WATCH",                    "Some pullback (-5% to 50% of typical dip) but not yet a deal."),
        ("",                     "HOLD",                     "Near all-time highs (less than -5% drawdown). Limited margin of safety."),
        ("── KEY METRICS ──",    "ROE %",                    ">20% Excellent | 15% Good | 10% Fair | <5% Weak"),
        ("",                     "Debt / Equity",            "<0.3 Excellent | <0.5 Good | <1.0 Fair | >2.0 Danger"),
        ("",                     "Current Ratio",            ">2.0 Excellent | >1.5 Good | <1.0 Red flag"),
        ("",                     "Interest Coverage",        ">10x Excellent | >5x Good | >3x Fair | <1.5x Danger"),
        ("",                     "FCF Yield %",              ">6% Excellent | >4% Good | >2% Fair | <0% Negative"),
        ("",                     "PEG Ratio",                "<1.0 Undervalued | <1.5 Good | >3.0 Expensive"),
        ("",                     "Revenue Growth %",         ">20% Strong | >10% Good | >5% Moderate | <0% Declining"),
        ("",                     "Gross Margin %",           ">50% Excellent | >40% Good | >30% Fair | <20% Thin"),
        ("── CYCLE METRICS ──",  "current_drawdown_pct",     "How far current price is below 252-bar ATH. More negative = deeper dip."),
        ("",                     "typical_drawdown",         "Average of all confirmed major dip pivot levels over full history."),
        ("",                     "lower_bound",              "Worst confirmed historical dip. Represents extreme fear level."),
        ("",                     "typical_profit",           "Average of all confirmed major rally pivot levels over full history."),
        ("",                     "upper_bound",              "Best confirmed historical rally level."),
        ("── COLOUR SCALE ──",   "Dark Green  #006400",      "Strong Buy / Excellent — best possible reading."),
        ("",                     "Forest Green  #228B22",    "Buy / Good — above average reading."),
        ("",                     "Gold  #FFCC00",            "Watch / Fair — acceptable but monitor closely."),
        ("",                     "Orange-Red  #FF4500",      "Hold / Weak — below acceptable threshold."),
        ("",                     "Firebrick  #B22222",       "Avoid / Poor — significant concern or red flag."),
        ("── SCORING WEIGHTS ──","Financial Health (40%)",   "Profitability 30% | Balance Sheet 25% | Growth 20% | Cash Flow 15% | Shareholder 10%"),
        ("",                     "Valuation Zone (35%)",     "Based on current_drawdown_pct vs typical_drawdown and lower_bound."),
        ("",                     "Momentum (25%)",           "Event count reliability 50% | Reward/risk ratio (typical_profit / |typical_drawdown|) 50%"),
    ]

    # Track max content length per column for autofit
    col_widths = [0, 0, 0]

    for ri, (a, b, c) in enumerate(rows, start=1):
        vals           = [a, b, c]
        is_top_header  = (a == "SECTION")
        is_section_div = a.startswith("──")

        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = mk_border()

            if is_top_header or is_section_div:
                # Header row AND section dividers: 1A1A1B background, White Bold
                cell.fill      = hdr_fill
                cell.font      = hdr_font
                cell.alignment = ctr_align if is_top_header else std_align
            else:
                # Regular data rows: white, normal text, no wrap
                cell.fill      = data_fill
                cell.font      = mk_font(bold=False, color=C_TEXT_BLACK, size=10)
                cell.alignment = std_align

            # Track max length for column autofit
            col_widths[ci - 1] = max(col_widths[ci - 1], len(str(val)) if val else 0)

        # Autofit row height — single line throughout (no wrap text)
        ws.row_dimensions[ri].height = 16

    # Autofit column widths based on content, minimum 18
    for ci, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = max(width + 4, 18)

    ws.freeze_panes = "A2"


# =============================================================================
# SAVE OUTPUT
# =============================================================================

def save_results(df: pd.DataFrame):
    # Ensure all columns present
    for col in ALL_COLS:
        if col not in df.columns:
            df[col] = None
    df_out = df[ALL_COLS].copy()

    # CSV
    df_out.to_csv(OUTPUT_CSV, index=False)
    logger.info(f"CSV saved: {OUTPUT_CSV}")

    # Excel — 3 sheets: Results, Summary, Legend
    try:
        with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="Results")
            wb = writer.book
            apply_sheet_formatting(wb["Results"], list(df_out.columns))
            add_summary_sheet(wb, df_out)
            add_legend_sheet(wb)
        logger.info(f"Excel saved: {OUTPUT_EXCEL}  (3 sheets: Results | Summary | Legend)")
    except Exception as e:
        logger.error(f"Excel save error: {e}", exc_info=True)


# =============================================================================
# MAIN
# =============================================================================

def main():
    logger.info("=" * 72)
    logger.info("  Major Cycle Analysis — Fundamentals + Cycle + Scoring Engine")
    logger.info(f"  Lookback={LOOKBACK_PERIOD} bars | PivotBars={PIVOT_BARS} | "
                f"PullbackThreshold={PULLBACK_THRESHOLD}% | ProfitThreshold={PROFIT_THRESHOLD}%")
    logger.info("=" * 72)

    # Create sample tickers.csv if missing
    if not os.path.exists(INPUT_CSV):
        logger.warning(f"{INPUT_CSV} not found — creating sample file")
        pd.DataFrame({"ticker": ["AAPL", "MSFT", "NVDA", "AMZN", "GOOGL",
                                  "JPM", "JNJ", "BRK-B", "V", "PG"]}).to_csv(
            INPUT_CSV, index=False)

    try:
        tdf = pd.read_csv(INPUT_CSV)
        if "ticker" not in tdf.columns:
            raise ValueError("Input CSV must contain a 'ticker' column")
        tickers = tdf["ticker"].dropna().str.strip().str.upper().unique().tolist()
        logger.info(f"Loaded {len(tickers)} tickers from {INPUT_CSV}")
    except Exception as e:
        logger.error(f"Could not load tickers: {e}")
        return

    start      = time.time()
    results_df = process_tickers(tickers)
    elapsed    = time.time() - start
    logger.info(f"Processing complete in {elapsed:.1f}s")

    save_results(results_df)

    success = results_df["current_close"].notna().sum()
    failed  = results_df["current_close"].isna().sum()
    logger.info(f"Done: {success} succeeded | {failed} failed | {len(tickers)} total")
    logger.info(f"Output files: {OUTPUT_EXCEL}  |  {OUTPUT_CSV}")

    # Print top 10 to console
    if "overall_rating" in results_df.columns:
        top10 = (
            results_df[results_df["overall_rating"].notna()]
            .sort_values("overall_rating", ascending=False)
            .head(10)[["ticker", "sector", "overall_rating", "overall_rating_label",
                        "financial_health_score", "valuation_zone",
                        "current_drawdown_pct", "typical_drawdown"]]
        )
        logger.info("\n\n" + "=" * 72)
        logger.info("  TOP 10 PICKS")
        logger.info("=" * 72)
        logger.info("\n" + top10.to_string(index=False))
        logger.info("=" * 72 + "\n")


if __name__ == "__main__":
    main()
